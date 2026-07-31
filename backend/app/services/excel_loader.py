"""
Universal Excel loader: supports .xlsx, .xlsm (openpyxl) and .xls (xlrd).
When openpyxl raises BadZipFile, falls back to xlrd for .xls files.
Handles large files via read_only mode and macro files via multiple strategies.
"""
from __future__ import annotations

import io
import logging
import re
import zipfile
from typing import Any

import openpyxl
import xlrd

logger = logging.getLogger(__name__)

# Excel formula error strings - treat as None when reading
_EXCEL_ERROR_STRINGS = frozenset({"#ref!", "#value!", "#div/0!", "#n/a", "#name?", "#num!", "#null!", "#getting_data"})


def _col_to_letter(col_1based: int) -> str:
    """Convert 1-based column index to Excel letter(s): 1->A, 26->Z, 27->AA."""
    result = ""
    n = col_1based
    while n > 0:
        n, r = divmod(n - 1, 26)
        result = chr(65 + r) + result
    return result


def _parse_cell_ref(cell_ref: str) -> tuple[int, int]:
    """Parse 'B2' -> (row_0based, col_0based)."""
    m = re.match(r"^([A-Za-z]+)(\d+)$", str(cell_ref).strip())
    if not m:
        raise ValueError(f"Invalid cell ref: {cell_ref}")
    col_letters, row_str = m.groups()
    row = int(row_str) - 1  # 1-based to 0-based
    col = 0
    for c in col_letters.upper():
        col = col * 26 + (ord(c) - ord("A") + 1)
    col -= 1  # 1-based to 0-based
    return row, col


def _normalize_cell_value(raw: Any) -> Any:
    """Treat Excel formula errors (#REF!, #VALUE!, etc.) as None."""
    if raw is None or raw == "":
        return None
    if isinstance(raw, str) and raw.strip().lower() in _EXCEL_ERROR_STRINGS:
        return None
    return raw


# =Sheet1!A1 or ='My Sheet'!A1 or =A1 (optional $)
_SIMPLE_CELL_FORMULA = re.compile(
    r"^=(?:'?([^'!]+)'?!)?(\$?[A-Za-z]+\$?\d+)\s*$",
    re.IGNORECASE,
)


def split_sheet_cell_ref(raw: str) -> tuple[str | None, str]:
    """
    Accept 'M65', 'INFO-8!M65', or \"'INFO-8'!M65\" → (sheet_or_None, 'M65').
    """
    s = str(raw).strip()
    if "!" in s:
        sheet_part, cell_part = s.rsplit("!", 1)
        sheet_part = sheet_part.strip().strip("'").strip('"')
        cell_part = cell_part.replace("$", "").strip()
        return (sheet_part or None), cell_part
    return None, s.replace("$", "").strip()


def get_cell_value(rows: list[list[Any]], cell_ref: str) -> Any:
    """Get cell value from rows (0-based row, col). Returns None if out of range or formula error."""
    try:
        _, bare = split_sheet_cell_ref(cell_ref)
        row_idx, col_idx = _parse_cell_ref(bare)
        if 0 <= row_idx < len(rows) and 0 <= col_idx < len(rows[row_idx]):
            raw = rows[row_idx][col_idx]
            return _normalize_cell_value(raw)
    except (ValueError, IndexError):
        pass
    return None


def _find_worksheet(wb: Any, sheet_name: str) -> Any | None:
    """Flexible sheet lookup: exact, case-insensitive, ignore spaces/hyphens."""
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    want = re.sub(r"[\s\-_]+", "", str(sheet_name).strip().lower())
    for name in wb.sheetnames:
        if re.sub(r"[\s\-_]+", "", name.strip().lower()) == want:
            return wb[name]
    # prefix match e.g. mapping "INFO-8" vs workbook "INFO-8 (EN)"
    for name in wb.sheetnames:
        base = name.split("(")[0].strip()
        if re.sub(r"[\s\-_]+", "", base.lower()) == want:
            return wb[name]
    return None


def read_mapped_cell(
    wb_values: Any,
    wb_formulas: Any | None,
    sheet_name: str,
    cell_ref: str,
    *,
    _depth: int = 0,
) -> tuple[Any, str | None]:
    """
    Read one mapped cell for extraction.

    - Prefers cached calculated value (data_only workbook).
    - If missing, follows simple `=A1` / `=Sheet!A1` formula chains from the formula workbook.
    - Returns (raw_value, number_format) so callers can scale percentage formats.
    """
    if _depth > 12:
        return None, None

    sheet_from_ref, bare_ref = split_sheet_cell_ref(cell_ref)
    use_sheet = sheet_from_ref or sheet_name
    bare_ref = bare_ref.replace("$", "")

    ws_val = _find_worksheet(wb_values, use_sheet)
    if ws_val is None:
        return None, None

    try:
        cell_val = ws_val[bare_ref]
    except Exception:
        return None, None

    number_format: str | None = None
    formula_text: str | None = None
    if wb_formulas is not None:
        ws_f = _find_worksheet(wb_formulas, use_sheet)
        if ws_f is not None:
            try:
                cell_f = ws_f[bare_ref]
                number_format = getattr(cell_f, "number_format", None)
                fv = cell_f.value
                if isinstance(fv, str) and fv.startswith("="):
                    formula_text = fv
            except Exception:
                pass

    raw = _normalize_cell_value(cell_val.value)
    if raw is not None and not (isinstance(raw, str) and raw.startswith("=")):
        return raw, number_format

    # Cached value missing — follow simple cell references only
    if formula_text:
        m = _SIMPLE_CELL_FORMULA.match(formula_text.strip())
        if m:
            ref_sheet = m.group(1)
            next_ref = m.group(2).replace("$", "")
            next_sheet = (ref_sheet.strip().strip("'") if ref_sheet else use_sheet)
            return read_mapped_cell(
                wb_values,
                wb_formulas,
                next_sheet,
                next_ref,
                _depth=_depth + 1,
            )

    return None, number_format


def open_workbooks_for_extraction(data: bytes) -> tuple[Any | None, Any | None]:
    """
    Open value (data_only) + formula workbooks for cell-level extraction.
    Returns (wb_values, wb_formulas). Either may be None on failure.
    """
    wb_values = None
    wb_formulas = None
    try:
        wb_values = openpyxl.load_workbook(
            io.BytesIO(data),
            data_only=True,
            keep_links=False,
            keep_vba=False,
        )
    except Exception as e:
        logger.warning("data_only workbook open failed: %s", e)

    try:
        wb_formulas = openpyxl.load_workbook(
            io.BytesIO(data),
            data_only=False,
            keep_links=False,
            keep_vba=False,
        )
    except Exception as e:
        logger.warning("formula workbook open failed: %s", e)

    return wb_values, wb_formulas


def close_workbooks(*workbooks: Any) -> None:
    for wb in workbooks:
        if wb is None:
            continue
        try:
            wb.close()
        except Exception:
            pass


def _load_with_openpyxl(data: bytes, **kwargs) -> list[tuple[str, list[list[Any]]]]:
    """Load with openpyxl using given options. Raises on failure."""
    wb = openpyxl.load_workbook(
        io.BytesIO(data),
        data_only=True,
        keep_links=False,  # Skip external links (can hang or fail)
        **kwargs,
    )
    result: list[tuple[str, list[list[Any]]]] = []
    for ws in wb.worksheets:
        rows: list[list[Any]] = []
        for r in ws.iter_rows(min_row=1, max_row=ws.max_row):
            rows.append([_normalize_cell_value(c.value) for c in r])
        result.append((ws.title, rows))
    return result


def _load_read_only(data: bytes) -> list[tuple[str, list[list[Any]]]]:
    """Load with openpyxl read_only mode (memory-efficient for large files)."""
    wb = openpyxl.load_workbook(
        io.BytesIO(data),
        read_only=True,
        data_only=True,
        keep_links=False,
        keep_vba=False,  # Skip VBA for read_only - faster and less memory
    )
    result: list[tuple[str, list[list[Any]]]] = []
    try:
        for ws in wb.worksheets:
            rows = [[_normalize_cell_value(c.value) for c in row] for row in ws.iter_rows()]
            result.append((ws.title, rows))
        return result
    finally:
        wb.close()


def load_excel_sheets(data: bytes) -> list[tuple[str, list[list[Any]]]]:
    """
    Load Excel file (xlsx, xlsm, xls) and return [(sheet_name, [[cell, ...], ...]), ...].
    Tries multiple strategies for complex files (macros, external links, large files).
    """
    last_error: Exception | None = None

    # Strategy 1: For large files (>2MB), try read_only first (memory-efficient)
    if len(data) > 2 * 1024 * 1024:
        try:
            logger.info("Trying read_only mode for large file (%s bytes)", len(data))
            return _load_read_only(data)
        except zipfile.BadZipFile:
            pass  # Not zip, try xlrd
        except Exception as e:
            last_error = e
            logger.warning("read_only failed: %s", e)

    # Strategy 2: Standard load - try keep_vba=False first (faster for macro files)
    for keep_vba in (False, True):
        try:
            return _load_with_openpyxl(data, keep_vba=keep_vba)
        except zipfile.BadZipFile:
            break  # Not a zip file, try xlrd
        except Exception as e:
            last_error = e
            logger.warning("openpyxl load (keep_vba=%s) failed: %s", keep_vba, e)
            continue

    # Strategy 3: read_only (for smaller files that failed standard load)
    if len(data) <= 2 * 1024 * 1024:
        try:
            return _load_read_only(data)
        except zipfile.BadZipFile:
            pass
        except Exception as e:
            last_error = e
            logger.warning("read_only fallback failed: %s", e)

    # Strategy 4: xlrd for .xls (Excel 97-2003)
    try:
        wb = xlrd.open_workbook(file_contents=data)
        result = []
        for i in range(wb.nsheets):
            sheet = wb.sheet_by_index(i)
            rows = []
            for r in range(sheet.nrows):
                row_vals = []
                for c in range(sheet.ncols):
                    val = _normalize_cell_value(sheet.cell_value(r, c))
                    row_vals.append(val)
                rows.append(row_vals)
            result.append((sheet.name, rows))
        return result
    except Exception as e:
        last_error = e
        logger.warning("xlrd fallback failed: %s", e)

    # All strategies failed - raise with clear message
    err_msg = str(last_error) if last_error else "Unknown error"
    raise ValueError(
        f"Could not read Excel file. Tried openpyxl and xlrd. Last error: {err_msg}"
    ) from last_error
